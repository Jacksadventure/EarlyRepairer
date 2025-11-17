#!/usr/bin/env python3
"""
Convert a SQLite database (.db) to an Excel workbook (.xlsx).
- Creates one sheet per table (excluding sqlite_internal tables)
- Auto-adjusts sheet names to Excel's 31-character limit and ensures uniqueness
- Optionally auto-sizes columns (best-effort)

Usage:
  python3 scripts/convert_db_to_excel.py --db single.db --out single.xlsx
"""
import argparse
import os
import sqlite3
import sys
from typing import List


def list_tables(conn: sqlite3.Connection) -> List[str]:
    cur = conn.cursor()
    cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name;"
    )
    return [r[0] for r in cur.fetchall()]


def list_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info('{table}')")
    return [row[1] for row in cur.fetchall()]


def safe_sheet_name(name: str, used: set) -> str:
    # Excel sheet name limit is 31 chars; also must be unique within workbook
    base = name[:31]
    candidate = base
    idx = 1
    while candidate in used:
        suffix = f"_{idx}"
        candidate = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        idx += 1
    used.add(candidate)
    return candidate


def main():
    parser = argparse.ArgumentParser(description="Convert SQLite .db to Excel .xlsx")
    parser.add_argument("--db", default="single.db", help="Path to SQLite database file")
    parser.add_argument("--out", default="single.xlsx", help="Output Excel file path")
    args = parser.parse_args()

    if not os.path.exists(args.db):
        print(f"Error: Database file not found: {args.db}", file=sys.stderr)
        sys.exit(1)

    # Try to use openpyxl directly (no pandas dependency)
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        print(
            "Error: openpyxl is required to write .xlsx files.\n"
            "Please install with: pip3 install --user openpyxl",
            file=sys.stderr,
        )
        sys.exit(2)

    with sqlite3.connect(args.db) as conn:
        tables = list_tables(conn)
        if not tables:
            print("No user tables found in database. Nothing to export.")
            sys.exit(0)

        wb = Workbook()
        # Remove the default sheet created by Workbook
        default_sheet = wb.active
        wb.remove(default_sheet)

        used = set()
        for t in tables:
            cols = list_columns(conn, t)
            sheet_name = safe_sheet_name(t, used)
            ws = wb.create_sheet(title=sheet_name)

            # Write header
            for c_idx, col in enumerate(cols, start=1):
                ws.cell(row=1, column=c_idx, value=col)

            # Stream rows to avoid high memory usage
            cur = conn.cursor()
            cur.execute(f"SELECT * FROM '{t}'")
            row_idx = 2
            for row in cur:
                for c_idx, val in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=c_idx, value=val)
                row_idx += 1

            # Best-effort auto-widths for columns (based on header only to stay fast)
            for c_idx, col in enumerate(cols, start=1):
                ws.column_dimensions[get_column_letter(c_idx)].width = max(10, min(len(str(col)) + 2, 80))

        wb.save(args.out)
        print(f"Exported {len(tables)} tables to {args.out}.")


if __name__ == "__main__":
    main()
